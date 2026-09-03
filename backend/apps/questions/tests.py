from django.contrib.auth import get_user_model
from django.core.exceptions import ValidationError
from django.core.files.uploadedfile import SimpleUploadedFile
from django.db import IntegrityError, transaction
from django.test import TestCase
from django.urls import reverse
from rest_framework import status
from rest_framework.test import APITestCase

from apps.taxonomy.models import ClientNeedsCategory, ClientNeedsSubcategory, NursingSystem, Topic

from .models import (
    MAX_QUESTION_IMAGE_BYTES,
    AnswerChoice,
    ClinicalJudgmentSkill,
    CognitiveLevel,
    Difficulty,
    Question,
    QuestionType,
)
from .services import QuestionNotGradeable, grade_submission

User = get_user_model()


def make_question(**overrides):
    """
    Shared test fixture builder: creates a minimal-but-valid taxonomy chain
    (NursingSystem -> Topic, ClientNeedsCategory -> ClientNeedsSubcategory)
    plus a Question referencing them, so individual tests don't each have to
    repeat this required-field boilerplate. **overrides lets a test replace
    any default (e.g. question_type=QuestionType.SATA) without redefining
    every other field.
    """

    # get_or_create, not create: the taxonomy rows are uniquely constrained
    # by name, so a test that builds a SECOND question would otherwise blow
    # up on a duplicate NursingSystem before it got anywhere near what it
    # was actually testing. Reusing the existing chain makes this helper
    # composable — call it as many times as a test needs, and every question
    # simply shares one taxonomy, which is what these tests want anyway.
    nursing_system, _ = NursingSystem.objects.get_or_create(name="Cardiovascular")
    topic, _ = Topic.objects.get_or_create(nursing_system=nursing_system, name="Heart Failure")
    category, _ = ClientNeedsCategory.objects.get_or_create(name="Physiological Adaptation")
    subcategory, _ = ClientNeedsSubcategory.objects.get_or_create(
        category=category, name="Illness Management"
    )

    # Only the fields Question actually requires (no null=True/blank=True in
    # models.py) are set here — subtopic, tags, rationale_incorrect,
    # reference, image, etc. are all optional and left unset, since a
    # minimal valid Question is exactly what most tests need.
    defaults = dict(
        question_type=QuestionType.MCQ,
        stem="A client with heart failure reports weight gain of 3 lbs in 2 days. What is the priority action?",
        difficulty=Difficulty.MEDIUM,
        nursing_system=nursing_system,
        topic=topic,
        nclex_client_needs_category=category,
        nclex_client_needs_subcategory=subcategory,
        clinical_judgment_skill=ClinicalJudgmentSkill.TAKE_ACTION,
        cognitive_level=CognitiveLevel.APPLY,
        rationale_correct="Rapid weight gain indicates fluid retention and should be reported immediately.",
    )
    defaults.update(overrides)
    return Question.objects.create(**defaults)


class QuestionTests(TestCase):
    def test_str_includes_type_and_stem(self):
        question = make_question()
        self.assertIn("MCQ", str(question))

    def test_missing_required_fk_raises(self):
        # nursing_system/topic/nclex_client_needs_category/
        # nclex_client_needs_subcategory are all required ForeignKeys (no
        # null=True in models.py) — omitting them entirely should fail at
        # the database level (NOT NULL constraint), confirming the schema
        # actually enforces "every question must be fully taxonomy-tagged"
        # rather than that constraint only existing in application code
        # that could be bypassed.
        with self.assertRaises(IntegrityError), transaction.atomic():
            Question.objects.create(
                question_type=QuestionType.MCQ,
                stem="Incomplete question",
                difficulty=Difficulty.EASY,
                clinical_judgment_skill=ClinicalJudgmentSkill.RECOGNIZE_CUES,
                cognitive_level=CognitiveLevel.REMEMBER,
                rationale_correct="n/a",
            )


class AnswerChoiceTests(TestCase):
    def test_ordered_by_display_order(self):
        # Deliberately creates "Second" (display_order=2) before "First"
        # (display_order=1) — if AnswerChoice.Meta.ordering weren't applied,
        # this query would return them in creation order ("Second" first)
        # instead of by display_order, so this test would only pass by
        # accident if the ordering weren't actually working.
        question = make_question(question_type=QuestionType.SATA)
        AnswerChoice.objects.create(question=question, choice_text="Second", display_order=2)
        AnswerChoice.objects.create(question=question, choice_text="First", display_order=1)

        # question.answer_choices is the related_name from
        # AnswerChoice.question — accessing it as a reverse relation
        # confirms both the ForeignKey wiring and the Meta ordering work
        # together correctly.
        ordered_texts = list(question.answer_choices.values_list("choice_text", flat=True))
        self.assertEqual(ordered_texts, ["First", "Second"])

    def test_sata_supports_multiple_correct_choices(self):
        # Confirms the schema has no constraint limiting is_correct=True to
        # a single row per question — essential for SATA (Select All That
        # Apply), which can have any number of correct choices, unlike MCQ.
        # See AnswerChoice's docstring in models.py: this rule isn't
        # enforced by the database at all, it's just that nothing here
        # prevents it.
        question = make_question(question_type=QuestionType.SATA)
        AnswerChoice.objects.create(
            question=question, choice_text="Correct 1", is_correct=True, display_order=1
        )
        AnswerChoice.objects.create(
            question=question, choice_text="Correct 2", is_correct=True, display_order=2
        )
        AnswerChoice.objects.create(question=question, choice_text="Wrong", is_correct=False, display_order=3)

        self.assertEqual(question.answer_choices.filter(is_correct=True).count(), 2)

    def test_rationale_is_optional_but_can_be_set_per_choice(self):
        # Each choice carries its own explanation (shown inline under that
        # option in the quiz UI) rather than relying on a single
        # question-level blob — confirms the field defaults to empty
        # (older/imported content may not have it yet) and holds a real
        # value when a choice does provide one.
        question = make_question()
        without_rationale = AnswerChoice.objects.create(question=question, choice_text="No explanation yet")
        with_rationale = AnswerChoice.objects.create(
            question=question,
            choice_text="Notify the provider",
            is_correct=True,
            rationale="Rapid weight gain indicates fluid retention and should be reported immediately.",
        )

        self.assertEqual(without_rationale.rationale, "")
        self.assertEqual(
            with_rationale.rationale,
            "Rapid weight gain indicates fluid retention and should be reported immediately.",
        )

    def test_question_rationale_correct_accepts_null(self):
        # rationale_correct was previously NOT NULL at the database level —
        # explicitly setting it to None would have raised an IntegrityError.
        # Now that AnswerChoice.rationale is the primary explanation
        # mechanism for MCQ/SATA/EMR questions, this must be genuinely
        # optional, not just "blank at the form layer."
        question = make_question(rationale_correct=None)
        self.assertIsNone(question.rationale_correct)


class GradeSubmissionTests(TestCase):
    """
    Direct tests of the grading rule, with no HTTP involved.

    Grading lives in services.py rather than in the view precisely so it can
    be exercised like this — and because Milestone 3's quiz engine and Phase
    2's analytics will call the same function. These tests are what stop the
    rule drifting when SATA partial credit arrives.
    """

    def setUp(self):
        self.question = make_question()
        self.right = AnswerChoice.objects.create(
            question=self.question, choice_text="Correct", is_correct=True, display_order=1
        )
        self.wrong = AnswerChoice.objects.create(
            question=self.question, choice_text="Wrong", is_correct=False, display_order=2
        )

    def test_selecting_the_correct_choice_is_correct(self):
        result = grade_submission(self.question, [self.right.id])
        self.assertTrue(result.is_correct)
        self.assertEqual(result.selected_ids, frozenset({self.right.id}))
        self.assertEqual(result.correct_ids, frozenset({self.right.id}))

    def test_selecting_a_distractor_is_incorrect(self):
        self.assertFalse(grade_submission(self.question, [self.wrong.id]).is_correct)

    def test_sata_requires_the_exact_set(self):
        question = make_question(
            question_type=QuestionType.SATA,
            stem="Select all that apply.",
            nursing_system=self.question.nursing_system,
            topic=self.question.topic,
            nclex_client_needs_category=self.question.nclex_client_needs_category,
            nclex_client_needs_subcategory=self.question.nclex_client_needs_subcategory,
        )
        a = AnswerChoice.objects.create(question=question, choice_text="A", is_correct=True, display_order=1)
        b = AnswerChoice.objects.create(question=question, choice_text="B", is_correct=True, display_order=2)
        c = AnswerChoice.objects.create(question=question, choice_text="C", is_correct=False, display_order=3)

        self.assertTrue(grade_submission(question, [a.id, b.id]).is_correct)
        # A partial selection is wrong under the current all-or-nothing
        # rule. This assertion is the one that will change when Phase 2
        # introduces partial credit — and it should change HERE, once.
        self.assertFalse(grade_submission(question, [a.id]).is_correct)
        # Superset: everything correct, plus a distractor.
        self.assertFalse(grade_submission(question, [a.id, b.id, c.id]).is_correct)

    def test_ids_from_another_question_are_discarded_not_counted(self):
        other = make_question(
            stem="A different question.",
            nursing_system=self.question.nursing_system,
            topic=self.question.topic,
            nclex_client_needs_category=self.question.nclex_client_needs_category,
            nclex_client_needs_subcategory=self.question.nclex_client_needs_subcategory,
        )
        foreign = AnswerChoice.objects.create(
            question=other, choice_text="Foreign", is_correct=True, display_order=1
        )

        result = grade_submission(self.question, [self.right.id, foreign.id])
        # Still correct: the foreign id is dropped rather than treated as an
        # extra selection that would spoil the exact-set match.
        self.assertTrue(result.is_correct)
        self.assertNotIn(foreign.id, result.selected_ids)

    def test_garbage_values_are_discarded_without_raising(self):
        # Defence in depth: the serializer normally rejects these before
        # grading is reached, but a future non-HTTP caller might not have
        # one, and grading must not explode.
        result = grade_submission(self.question, ["not-a-uuid", None, 42, self.right.id])
        self.assertTrue(result.is_correct)

    def test_question_with_no_correct_choice_raises(self):
        # The bug this guards: with no correct choices the expected set is
        # empty, so a naive set comparison reports an empty submission as
        # CORRECT — telling a student they were right about a question that
        # has no right answer.
        broken = make_question(
            stem="Broken question with no correct answer.",
            nursing_system=self.question.nursing_system,
            topic=self.question.topic,
            nclex_client_needs_category=self.question.nclex_client_needs_category,
            nclex_client_needs_subcategory=self.question.nclex_client_needs_subcategory,
        )
        AnswerChoice.objects.create(question=broken, choice_text="A", is_correct=False, display_order=1)

        with self.assertRaises(QuestionNotGradeable):
            grade_submission(broken, [])


class QuestionListAPITests(APITestCase):
    def setUp(self):
        self.user = User.objects.create_user(
            email="student@example.com", password="a-strong-password-123", is_active=True
        )
        self.question = make_question()
        self.choice = AnswerChoice.objects.create(
            question=self.question,
            choice_text="Notify the provider",
            is_correct=True,
            display_order=1,
            rationale="Rapid weight gain signals fluid retention.",
        )

    def test_requires_authentication(self):
        self.assertEqual(self.client.get(reverse("question-list")).status_code, status.HTTP_401_UNAUTHORIZED)

    def test_response_is_paginated(self):
        # The endpoint used to return a bare list of every active question.
        # At the specced 4,000+ questions that response ran to megabytes, so
        # the paginated envelope is load-bearing, not cosmetic.
        self.client.force_authenticate(self.user)
        response = self.client.get(reverse("question-list"))

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertIn("results", response.data)
        self.assertIn("count", response.data)

    def test_list_never_exposes_the_answer_key(self):
        # The single most important assertion in this file. If
        # PublicAnswerChoiceSerializer ever starts including is_correct or
        # rationale, every quiz becomes trivially cheatable by reading the
        # question list — and nothing else in the suite would notice.
        self.client.force_authenticate(self.user)
        response = self.client.get(reverse("question-list"))

        choice = response.data["results"][0]["answer_choices"][0]
        self.assertNotIn("is_correct", choice)
        self.assertNotIn("rationale", choice)
        self.assertIn("choice_text", choice)

    def test_inactive_questions_are_excluded(self):
        self.question.is_active = False
        self.question.save(update_fields=["is_active"])
        self.client.force_authenticate(self.user)

        self.assertEqual(self.client.get(reverse("question-list")).data["count"], 0)


class QuestionSubmitAPITests(APITestCase):
    def setUp(self):
        self.user = User.objects.create_user(
            email="student@example.com", password="a-strong-password-123", is_active=True
        )
        self.client.force_authenticate(self.user)
        self.question = make_question()
        self.right = AnswerChoice.objects.create(
            question=self.question,
            choice_text="Right",
            is_correct=True,
            display_order=1,
            rationale="Because.",
        )
        self.wrong = AnswerChoice.objects.create(
            question=self.question, choice_text="Wrong", is_correct=False, display_order=2, rationale="Nope."
        )
        self.url = reverse("question-submit", args=[self.question.pk])

    def submit(self, payload):
        return self.client.post(self.url, payload, format="json")

    def test_requires_authentication(self):
        self.client.force_authenticate(None)
        self.assertEqual(
            self.submit({"selected_choice_ids": [str(self.right.id)]}).status_code,
            status.HTTP_401_UNAUTHORIZED,
        )

    def test_correct_answer_is_graded_and_reveals_the_key(self):
        response = self.submit({"selected_choice_ids": [str(self.right.id)]})

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertTrue(response.data["is_correct"])
        # Revealing the key is the intended behaviour AFTER answering — the
        # student needs the rationale to learn from the question.
        keyed = {c["id"]: c for c in response.data["choices"]}
        self.assertTrue(keyed[str(self.right.id)]["is_correct"])
        self.assertEqual(keyed[str(self.right.id)]["rationale"], "Because.")

    def test_wrong_answer_is_graded_as_incorrect(self):
        self.assertFalse(self.submit({"selected_choice_ids": [str(self.wrong.id)]}).data["is_correct"])

    def test_integer_payload_returns_400_not_500(self):
        # The exact crash this endpoint used to have: a bare number hit
        # `for i in request.data.get(...)` and raised TypeError out of the
        # view, producing a 500.
        self.assertEqual(self.submit({"selected_choice_ids": 5}).status_code, status.HTTP_400_BAD_REQUEST)

    def test_object_payload_returns_400_not_silently_graded(self):
        # Worse than the crash: a dict used to iterate its KEYS and grade
        # against them, so a malformed request was silently accepted.
        self.assertEqual(
            self.submit({"selected_choice_ids": {"a": 1}}).status_code, status.HTTP_400_BAD_REQUEST
        )

    def test_string_payload_returns_400(self):
        self.assertEqual(self.submit({"selected_choice_ids": "abc"}).status_code, status.HTTP_400_BAD_REQUEST)

    def test_missing_field_returns_400(self):
        self.assertEqual(self.submit({}).status_code, status.HTTP_400_BAD_REQUEST)

    def test_empty_selection_is_rejected_and_reveals_nothing(self):
        # An empty submission used to return 200 with the full answer key,
        # which made this endpoint a plain "give me the answers" call for
        # any question the student chose to skip.
        response = self.submit({"selected_choice_ids": []})

        self.assertEqual(response.status_code, status.HTTP_400_BAD_REQUEST)
        self.assertNotIn("choices", response.data)

    def test_non_uuid_id_returns_400(self):
        self.assertEqual(
            self.submit({"selected_choice_ids": ["not-a-uuid"]}).status_code, status.HTTP_400_BAD_REQUEST
        )

    def test_unknown_but_valid_uuid_is_graded_as_a_wrong_answer(self):
        # A well-formed id that isn't one of this question's choices is
        # dropped, leaving an empty selection — which cannot match the
        # answer key, so it grades as incorrect rather than erroring.
        response = self.submit({"selected_choice_ids": ["00000000-0000-0000-0000-000000000000"]})

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertFalse(response.data["is_correct"])

    def test_inactive_question_returns_404(self):
        self.question.is_active = False
        self.question.save(update_fields=["is_active"])

        self.assertEqual(
            self.submit({"selected_choice_ids": [str(self.right.id)]}).status_code,
            status.HTTP_404_NOT_FOUND,
        )

    def test_question_with_no_correct_choice_returns_409(self):
        self.right.is_correct = False
        self.right.save(update_fields=["is_correct"])

        response = self.submit({"selected_choice_ids": [str(self.right.id)]})

        # 409, not 200-with-is_correct-false: the question is broken
        # content, and the student must not be told they answered it wrong.
        self.assertEqual(response.status_code, status.HTTP_409_CONFLICT)
        self.assertNotIn("choices", response.data)


class ExternalIdTests(TestCase):
    """
    external_id is the natural key the importer matches on. Its uniqueness
    is what makes re-importing safe; its nullability is what keeps
    hand-authored admin questions possible.
    """

    def test_external_id_is_unique(self):
        first = make_question(external_id="NW-MCQ-001")
        with self.assertRaises(IntegrityError), transaction.atomic():
            make_question(
                external_id="NW-MCQ-001",
                stem="A different stem entirely.",
                nursing_system=first.nursing_system,
                topic=first.topic,
                nclex_client_needs_category=first.nclex_client_needs_category,
                nclex_client_needs_subcategory=first.nclex_client_needs_subcategory,
            )

    def test_multiple_questions_may_have_no_external_id(self):
        # NULLs are distinct under a unique constraint, so any number of
        # admin-authored questions can coexist without one.
        first = make_question()
        make_question(
            stem="Another hand-written question.",
            nursing_system=first.nursing_system,
            topic=first.topic,
            nclex_client_needs_category=first.nclex_client_needs_category,
            nclex_client_needs_subcategory=first.nclex_client_needs_subcategory,
        )
        self.assertEqual(Question.objects.filter(external_id__isnull=True).count(), 2)


class QuestionImageValidationTests(TestCase):
    """
    Question.image is a FileField with no content-type checking of its own.
    These validators are what stop an executable file being stored as a
    "question image" — which becomes stored XSS on the platform's own origin
    the moment media is served from a real host (see MEDIA_ROOT's note in
    settings; production still needs an object store in front of it).
    """

    def test_svg_upload_is_rejected(self):
        # SVG is the important case: it is an XML document that can carry
        # <script>, so browsers execute it rather than merely decoding it.
        question = make_question()
        question.image = SimpleUploadedFile(
            "payload.svg", b"<svg xmlns='http://www.w3.org/2000/svg'><script>alert(1)</script></svg>"
        )
        with self.assertRaises(ValidationError):
            question.full_clean()

    def test_html_upload_is_rejected(self):
        question = make_question()
        question.image = SimpleUploadedFile("payload.html", b"<script>alert(1)</script>")
        with self.assertRaises(ValidationError):
            question.full_clean()

    def test_oversized_image_is_rejected(self):
        question = make_question()
        question.image = SimpleUploadedFile(
            "huge.png", b"x" * (MAX_QUESTION_IMAGE_BYTES + 1), content_type="image/png"
        )
        with self.assertRaises(ValidationError):
            question.full_clean()

    def test_ordinary_png_is_accepted(self):
        question = make_question()
        question.image = SimpleUploadedFile("diagram.png", b"fake png bytes", content_type="image/png")
        # Must not raise — the validators have to leave legitimate content
        # alone or they simply block the feature.
        question.full_clean()


class PaginationContractTests(APITestCase):
    """
    Pins the exact envelope the frontend's listQuestions() depends on.

    That client pages by explicit ?page=N and stops when `next` is null,
    accumulating `results` into one flat array (see
    frontend/src/lib/api/questions.ts). If any of those three keys were
    renamed or the traversal terminated differently, the quiz setup page
    would silently receive a partial question bank — it builds its filter
    dropdowns and match counts from whatever it gets back, so a truncated
    list looks like working software rather than an error.
    """

    def setUp(self):
        self.user = User.objects.create_user(
            email="student@example.com", password="a-strong-password-123", is_active=True
        )
        self.client.force_authenticate(self.user)
        # Deliberately more than one page: DefaultPagination.page_size is 50,
        # so 55 questions forces the multi-page path the client relies on.
        # A single-page fixture would pass even if traversal were broken.
        for index in range(55):
            make_question(stem=f"Question number {index}.", external_id=f"NW-PAGE-{index:03d}")

    def test_envelope_exposes_the_keys_the_client_reads(self):
        data = self.client.get(reverse("question-list")).data
        for key in ("count", "next", "previous", "results"):
            self.assertIn(key, data)

    def test_paging_by_explicit_page_number_returns_every_question(self):
        # Mirrors the client loop exactly, including its termination
        # condition, rather than asserting on page contents directly.
        collected = []
        page = 1
        while page <= 200:
            data = self.client.get(reverse("question-list"), {"page": page}).data
            collected.extend(data["results"])
            if not data["next"]:
                break
            page += 1

        self.assertEqual(len(collected), 55)
        self.assertEqual(data["count"], 55)
        # No duplicates across pages — an off-by-one in either the client's
        # or the server's paging would show up here.
        self.assertEqual(len({q["id"] for q in collected}), 55)

    def test_last_page_reports_next_as_null(self):
        # The client's only stop signal. If this were ever an empty string
        # or an absent key, the loop would run to its MAX_PAGES ceiling on
        # every single page load.
        data = self.client.get(reverse("question-list"), {"page": 2}).data
        self.assertIsNone(data["next"])


# --- import_ngn_item_bank -----------------------------------------------

import tempfile

import openpyxl
from django.core.management import call_command

from apps.taxonomy.models import CaseStudy, Domain

from .models import MatrixCell, MatrixColumn, MatrixRow


def _write_sheet(wb, name, headers, rows):
    ws = wb.create_sheet(name)
    ws.append(headers)
    for row in rows:
        ws.append([row.get(h) for h in headers])


def _build_workbook():
    """
    A small, self-contained fixture exercising one question of every type
    plus a 2-item case study — not the real client file, just enough to
    verify import_ngn_item_bank's actual behavior end to end.
    """
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    im_headers = [
        "Question_ID",
        "Item_Type",
        "Domain",
        "Body_System",
        "Topic",
        "Subtopic",
        "Difficulty",
        "Cognitive_Level",
        "Client_Needs_Category",
        "Client_Needs_Subcategory",
        "Clinical_Judgment_Skill_Primary",
        "Clinical_Judgment_Skill_Secondary",
        "Tags",
        "Scenario",
        "Stem",
        "Clinical_Tip",
        "Reference_Key",
    ]
    common = dict(
        Domain="Adult Health",
        Body_System="Cardiovascular",
        Topic="Heart Failure",
        Subtopic=None,
        Difficulty="Moderate",
        Cognitive_Level="Apply",
        Client_Needs_Category="Physiological Integrity",
        Client_Needs_Subcategory="Reduction of Risk Potential",
        Clinical_Judgment_Skill_Primary="Take Action",
        Clinical_Judgment_Skill_Secondary=None,
        Tags="test, fixture",
        Scenario=None,
        Clinical_Tip="Tip.",
        Reference_Key=None,
    )
    im_rows = [
        dict(Question_ID="MCQ-1", Item_Type="MCQ", Stem="Which is correct?", **common),
        dict(Question_ID="MTX-1", Item_Type="Matrix/Grid", Stem="Expected or unexpected?", **common),
        dict(Question_ID="BT-1", Item_Type="Bow-Tie", Stem="Complete the bow-tie.", **common),
        dict(Question_ID="CLZ-1", Item_Type="Cloze", Stem="The nurse should [dropdown 1].", **common),
        dict(
            Question_ID="EHS-1",
            Item_Type="Hot Spot",
            Stem="Highlight the finding: client is diaphoretic and pale.",
            **{**common, "Scenario": None},
        ),
        dict(Question_ID="DND-1", Item_Type="Drag-and-Drop (Sequencing)", Stem="Order the steps.", **common),
        dict(Question_ID="CASE-1", Item_Type="NGN Case Study", Stem="See NGN_Components", **common),
    ]
    _write_sheet(wb, "Item_Master", im_headers, im_rows)

    ao_headers = ["Question_ID", "Option_ID", "Option_Text", "Is_Correct", "Rationale"]
    ao_rows = [
        dict(
            Question_ID="MCQ-1", Option_ID="A", Option_Text="Right", Is_Correct="TRUE", Rationale="Because."
        ),
        dict(
            Question_ID="MCQ-1", Option_ID="B", Option_Text="Wrong", Is_Correct="FALSE", Rationale="Not this."
        ),
        dict(
            Question_ID="MTX-1",
            Option_ID="Finding A",
            Option_Text="Expected",
            Is_Correct="TRUE",
            Rationale="Normal.",
        ),
        dict(
            Question_ID="MTX-1",
            Option_ID="Finding B",
            Option_Text="Unexpected",
            Is_Correct="TRUE",
            Rationale="Not normal.",
        ),
        dict(
            Question_ID="BT-1",
            Option_ID="Action 1",
            Option_Text="Do X",
            Is_Correct="TRUE",
            Rationale="X helps.",
        ),
        dict(
            Question_ID="BT-1",
            Option_ID="Action 2",
            Option_Text="Do Y",
            Is_Correct="FALSE",
            Rationale="Y doesn't.",
        ),
        dict(
            Question_ID="BT-1",
            Option_ID="Condition A",
            Option_Text="Condition X",
            Is_Correct="TRUE",
            Rationale="Matches.",
        ),
        dict(
            Question_ID="BT-1",
            Option_ID="Assessment i",
            Option_Text="Vitals",
            Is_Correct="TRUE",
            Rationale="Watch vitals.",
        ),
        dict(
            Question_ID="CLZ-1",
            Option_ID="Option 1",
            Option_Text="call the provider",
            Is_Correct="TRUE",
            Rationale="Right.",
        ),
        dict(
            Question_ID="CLZ-1",
            Option_ID="Option 2",
            Option_Text="do nothing",
            Is_Correct="FALSE",
            Rationale="Wrong.",
        ),
        dict(
            Question_ID="EHS-1",
            Option_ID="Highlight 1",
            Option_Text="diaphoretic and pale",
            Is_Correct="TRUE",
            Rationale="Concerning.",
        ),
        dict(
            Question_ID="DND-1",
            Option_ID="Step 1",
            Option_Text="First",
            Is_Correct="TRUE",
            Rationale="Goes first.",
        ),
        dict(
            Question_ID="DND-1",
            Option_ID="Step 2",
            Option_Text="Second",
            Is_Correct="TRUE",
            Rationale="Goes second.",
        ),
    ]
    _write_sheet(wb, "Answer_Options", ao_headers, ao_rows)

    nc_headers = [
        "Case_ID",
        "Domain",
        "Body_System",
        "Topic/Subtopic",
        "Overall_Difficulty",
        "Reference_Key",
        "Case_Presentation_Hour0",
    ]
    _write_sheet(
        wb,
        "NGN_Cases",
        nc_headers,
        [
            dict(
                Case_ID="CASE-1",
                Domain="Adult Health",
                Body_System="Cardiovascular",
                **{"Topic/Subtopic": "Heart Failure"},
                Overall_Difficulty="Moderate",
                Reference_Key=None,
                Case_Presentation_Hour0="Hour 0 presentation.",
            ),
        ],
    )

    ncomp_headers = [
        "Case_ID",
        "Item_No",
        "Clinical_Judgment_Step",
        "Item_Type",
        "Item_Subtype_Detail",
        "Difficulty",
        "Cognitive_Level",
        "Domain",
        "Body_System",
        "Topic",
        "Subtopic",
        "Client_Needs_Category",
        "Client_Needs_Subcategory",
        "Reference_Key",
        "Updated_Exhibit",
        "Stem",
        "Correct_Answer",
        "Rationale",
        "Clinical_Tip",
    ]
    _write_sheet(
        wb,
        "NGN_Components",
        ncomp_headers,
        [
            dict(
                Case_ID="CASE-1",
                Item_No=1,
                Clinical_Judgment_Step="Recognize Cues",
                Item_Type="MCQ",
                Item_Subtype_Detail=None,
                Difficulty="Moderate",
                Cognitive_Level="Analyze",
                Domain=None,
                Body_System=None,
                Topic=None,
                Subtopic=None,
                Client_Needs_Category="Physiological Integrity",
                Client_Needs_Subcategory="Reduction of Risk Potential",
                Reference_Key=None,
                Updated_Exhibit="Hour 0",
                Stem="A) foo B) bar",
                Correct_Answer="A",
                Rationale="Because A.",
                Clinical_Tip=None,
            ),
            dict(
                Case_ID="CASE-1",
                Item_No=2,
                Clinical_Judgment_Step="Take Action",
                Item_Type="SATA",
                Item_Subtype_Detail=None,
                Difficulty="Moderate",
                Cognitive_Level="Apply",
                Domain=None,
                Body_System=None,
                Topic=None,
                Subtopic=None,
                Client_Needs_Category="Physiological Integrity",
                Client_Needs_Subcategory="Reduction of Risk Potential",
                Reference_Key=None,
                Updated_Exhibit="Hour 1",
                Stem="A) foo B) bar C) baz",
                Correct_Answer="A, C",
                Rationale="Because.",
                Clinical_Tip=None,
            ),
        ],
    )

    _write_sheet(wb, "References", ["Reference_Key", "Full_Citation"], [])

    return wb


class ImportNgnItemBankTests(TestCase):
    def setUp(self):
        nursing_system, _ = NursingSystem.objects.get_or_create(name="Cardiovascular")
        Topic.objects.get_or_create(nursing_system=nursing_system, name="Heart Failure")
        Domain.objects.get_or_create(name="Adult Health")
        category, _ = ClientNeedsCategory.objects.get_or_create(name="Physiological Integrity")
        ClientNeedsSubcategory.objects.get_or_create(category=category, name="Reduction of Risk Potential")

    def _run(self, wb, **kwargs):
        with tempfile.NamedTemporaryFile(suffix=".xlsx") as f:
            wb.save(f.name)
            call_command("import_ngn_item_bank", file=f.name, **kwargs)

    def test_imports_one_of_every_type_plus_a_case_study(self):
        self._run(_build_workbook())

        self.assertEqual(Question.objects.filter(external_id="MCQ-1").count(), 1)
        self.assertEqual(Question.objects.filter(external_id="MTX-1").count(), 1)
        self.assertEqual(Question.objects.filter(external_id="BT-1").count(), 1)
        self.assertEqual(Question.objects.filter(external_id="CLZ-1").count(), 1)
        self.assertEqual(Question.objects.filter(external_id="EHS-1").count(), 1)
        self.assertEqual(Question.objects.filter(external_id="DND-1").count(), 1)
        # The case-study summary row in Item_Master must NOT become its own
        # Question — only its 2 linked NGN_Components items should.
        self.assertFalse(Question.objects.filter(external_id="CASE-1").exists())
        self.assertEqual(Question.objects.filter(external_id__startswith="CASE-1-item-").count(), 2)

    def test_matrix_synthesizes_the_false_cell_for_every_row(self):
        self._run(_build_workbook())
        question = Question.objects.get(external_id="MTX-1")
        # 2 rows x 2 columns = 4 cells, even though the sheet only stated
        # ONE correct column per row — the other cell must be synthesized.
        self.assertEqual(MatrixCell.objects.filter(row__question=question).count(), 4)
        correct = MatrixCell.objects.filter(row__question=question, is_correct=True)
        self.assertEqual(correct.count(), 2)

    def test_bowtie_splits_options_into_the_right_sections(self):
        self._run(_build_workbook())
        question = Question.objects.get(external_id="BT-1")
        sections = {o.section for o in question.bowtie_options.all()}
        self.assertEqual(sections, {"ACTION", "CONDITION", "ASSESSMENT"})

    def test_case_study_items_get_correct_sequence_and_ngn_type(self):
        self._run(_build_workbook())
        case = CaseStudy.objects.get(external_id="CASE-1")
        items = list(Question.objects.filter(case_study=case).order_by("case_study_sequence"))
        self.assertEqual([i.case_study_sequence for i in items], [1, 2])
        self.assertEqual([i.question_type for i in items], ["NGN_CASE", "NGN_CASE"])
        self.assertEqual([i.ngn_type for i in items], ["MCQ", "SATA"])
        # Item 1 has no Domain/Body_System override in the fixture — must
        # inherit from the NGN_Cases row rather than being left null.
        self.assertEqual(items[0].domain.name, "Adult Health")
        self.assertEqual(items[0].nursing_system.name, "Cardiovascular")

    def test_hotspot_target_not_found_in_stem_is_rejected_not_imported(self):
        # Regression test for the exact bug caught while building this
        # command: EHS-002's target text didn't match its own passage
        # verbatim. RowError is caught per-row inside handle() and reported
        # to stdout rather than propagating — so the observable behavior a
        # caller can assert on is "the row was never created", not an
        # exception reaching call_command.
        wb = _build_workbook()
        ws = wb["Answer_Options"]
        for row in ws.iter_rows(min_row=2):
            if row[0].value == "EHS-1":
                row[2].value = "this text is not in the stem at all"

        self._run(wb)
        self.assertFalse(Question.objects.filter(external_id="EHS-1").exists())

    def test_matrix_with_more_than_two_columns_is_accepted(self):
        # The importer's column set is inferred from the union of each
        # row's own correct-answer label, so a 3rd row naming a 3rd
        # distinct column ("Neither") both adds a new matrix row AND
        # grows the column count to 3 — this now matches the admin
        # dashboard's builder, which has never capped column count.
        wb = _build_workbook()
        ws = wb["Answer_Options"]
        ws.append(["MTX-1", "Finding C", "Neither", "TRUE", "Ambiguous."])
        self._run(wb)

        question = Question.objects.get(external_id="MTX-1")
        self.assertEqual(MatrixColumn.objects.filter(question=question).count(), 3)
        self.assertEqual(MatrixRow.objects.filter(question=question).count(), 3)
        # 3 rows x 3 columns = 9 cells, synthesized the same way the
        # 2-column case already was — one correct cell per row, the rest
        # false.
        self.assertEqual(MatrixCell.objects.filter(row__question=question).count(), 9)
        self.assertEqual(MatrixCell.objects.filter(row__question=question, is_correct=True).count(), 3)

    def test_matrix_with_fewer_than_two_columns_is_rejected(self):
        # Every row in this fixture names the SAME column as its correct
        # answer, so the union of correct-answer labels never grows past 1.
        wb = _build_workbook()
        ws = wb["Answer_Options"]
        for row in ws.iter_rows(min_row=2):
            if row[0].value == "MTX-1":
                row[2].value = "Expected"
        self._run(wb)
        self.assertFalse(Question.objects.filter(external_id="MTX-1").exists())

    def test_rerunning_is_idempotent(self):
        wb = _build_workbook()
        self._run(wb)
        count_before = Question.objects.count()
        self._run(wb)
        self.assertEqual(Question.objects.count(), count_before)

    def test_dry_run_writes_nothing(self):
        self._run(_build_workbook(), dry_run=True)
        self.assertEqual(Question.objects.count(), 0)
        self.assertEqual(CaseStudy.objects.count(), 0)


class AuthoringStructureValidationTests(TestCase):
    """
    Direct unit tests of apps.questions.authoring.validate_structure — one
    per structural rule, independent of the admin serializer/HTTP layer
    that also exercises it. Each of these corresponds to a question that
    would otherwise save cleanly and then blow up at quiz time with
    QuestionNotGradeable (apps.questions.services) — that's the whole
    reason this validation exists.
    """

    def setUp(self):
        from .authoring import validate_structure

        self.validate_structure = validate_structure

    def _validate(
        self, effective_type, structure, stem="Stem text.", clinical_scenario=None, case_scenario=None
    ):
        self.validate_structure(
            effective_type=effective_type,
            stem=stem,
            clinical_scenario=clinical_scenario,
            case_scenario=case_scenario,
            structure=structure,
        )

    # --- MCQ / SATA / EMR ---

    def test_mcq_with_two_correct_choices_is_rejected(self):
        from .authoring import StructureError

        choices = [
            {"choice_text": "A", "is_correct": True, "display_order": 0, "rationale": ""},
            {"choice_text": "B", "is_correct": True, "display_order": 1, "rationale": ""},
        ]
        with self.assertRaises(StructureError):
            self._validate(QuestionType.MCQ, {"answer_choices": choices})

    def test_sata_with_zero_correct_choices_is_rejected(self):
        from .authoring import StructureError

        choices = [
            {"choice_text": "A", "is_correct": False, "display_order": 0, "rationale": ""},
            {"choice_text": "B", "is_correct": False, "display_order": 1, "rationale": ""},
        ]
        with self.assertRaises(StructureError):
            self._validate(QuestionType.SATA, {"answer_choices": choices})

    def test_sata_with_one_choice_is_rejected(self):
        from .authoring import StructureError

        choices = [{"choice_text": "A", "is_correct": True, "display_order": 0, "rationale": ""}]
        with self.assertRaises(StructureError):
            self._validate(QuestionType.SATA, {"answer_choices": choices})

    def test_duplicate_display_order_is_rejected(self):
        from .authoring import StructureError

        choices = [
            {"choice_text": "A", "is_correct": True, "display_order": 0, "rationale": ""},
            {"choice_text": "B", "is_correct": False, "display_order": 0, "rationale": ""},
        ]
        with self.assertRaises(StructureError):
            self._validate(QuestionType.MCQ, {"answer_choices": choices})

    def test_valid_mcq_passes(self):
        choices = [
            {"choice_text": "A", "is_correct": True, "display_order": 0, "rationale": ""},
            {"choice_text": "B", "is_correct": False, "display_order": 1, "rationale": ""},
        ]
        self._validate(QuestionType.MCQ, {"answer_choices": choices})  # no exception

    # --- MATRIX ---

    def _matrix_structure(self, **overrides):
        columns = [
            {"key": "c0", "text": "Expected", "display_order": 0},
            {"key": "c1", "text": "Unexpected", "display_order": 1},
        ]
        rows = [
            {
                "key": "r0",
                "text": "BP 88/54",
                "display_order": 0,
                "cells": [
                    {"column_key": "c0", "is_correct": False, "rationale": ""},
                    {"column_key": "c1", "is_correct": True, "rationale": ""},
                ],
            }
        ]
        structure = {"matrix_columns": columns, "matrix_rows": rows}
        structure.update(overrides)
        return structure

    def test_matrix_row_with_no_correct_cell_is_rejected(self):
        from .authoring import StructureError

        structure = self._matrix_structure()
        structure["matrix_rows"][0]["cells"][1]["is_correct"] = False
        with self.assertRaises(StructureError):
            self._validate(QuestionType.MATRIX, structure)

    def test_matrix_row_missing_a_cell_for_one_column_is_rejected(self):
        from .authoring import StructureError

        structure = self._matrix_structure()
        structure["matrix_rows"][0]["cells"].pop()
        with self.assertRaises(StructureError):
            self._validate(QuestionType.MATRIX, structure)

    def test_matrix_cell_with_unresolvable_column_key_is_rejected(self):
        from .authoring import StructureError

        structure = self._matrix_structure()
        structure["matrix_rows"][0]["cells"][0]["column_key"] = "does-not-exist"
        with self.assertRaises(StructureError):
            self._validate(QuestionType.MATRIX, structure)

    def test_matrix_with_fewer_than_two_columns_is_rejected(self):
        from .authoring import StructureError

        structure = self._matrix_structure()
        structure["matrix_columns"].pop()
        structure["matrix_rows"][0]["cells"].pop()
        with self.assertRaises(StructureError):
            self._validate(QuestionType.MATRIX, structure)

    def test_valid_matrix_passes(self):
        self._validate(QuestionType.MATRIX, self._matrix_structure())  # no exception

    # --- BOWTIE ---

    def _bowtie_structure(self, **overrides):
        options = [
            {
                "section": "ASSESSMENT",
                "option_text": "Crackles",
                "is_correct": True,
                "display_order": 0,
                "rationale": "",
            },
            {
                "section": "CONDITION",
                "option_text": "Fluid overload",
                "is_correct": True,
                "display_order": 0,
                "rationale": "",
            },
            {
                "section": "ACTION",
                "option_text": "Give furosemide",
                "is_correct": True,
                "display_order": 0,
                "rationale": "",
            },
        ]
        structure = {"bowtie_options": options}
        structure.update(overrides)
        return structure

    def test_bowtie_section_with_no_correct_option_is_rejected(self):
        from .authoring import StructureError

        structure = self._bowtie_structure()
        structure["bowtie_options"][0]["is_correct"] = False
        with self.assertRaises(StructureError):
            self._validate(QuestionType.BOWTIE, structure)

    def test_bowtie_missing_a_section_entirely_is_rejected(self):
        from .authoring import StructureError

        structure = self._bowtie_structure()
        structure["bowtie_options"] = [o for o in structure["bowtie_options"] if o["section"] != "ACTION"]
        with self.assertRaises(StructureError):
            self._validate(QuestionType.BOWTIE, structure)

    def test_valid_bowtie_passes(self):
        self._validate(QuestionType.BOWTIE, self._bowtie_structure())  # no exception

    # --- CLOZE ---

    def _cloze_structure(self, stem="The nurse should assess the client's [dropdown 1] first."):
        blanks = [
            {
                "blank_key": "dropdown 1",
                "display_order": 0,
                "options": [
                    {"option_text": "airway", "is_correct": True, "rationale": ""},
                    {"option_text": "temperature", "is_correct": False, "rationale": ""},
                ],
            }
        ]
        return {"cloze_blanks": blanks}, stem

    def test_cloze_blank_key_with_no_stem_placeholder_is_rejected(self):
        from .authoring import StructureError

        structure, _ = self._cloze_structure()
        with self.assertRaises(StructureError):
            self._validate(QuestionType.CLOZE, structure, stem="No placeholders here.")

    def test_cloze_stem_placeholder_with_no_matching_blank_is_rejected(self):
        from .authoring import StructureError

        structure, stem = self._cloze_structure(
            stem="The nurse should assess the client's [dropdown 1] and [dropdown 2]."
        )
        with self.assertRaises(StructureError):
            self._validate(QuestionType.CLOZE, structure, stem=stem)

    def test_cloze_blank_with_zero_correct_options_is_rejected(self):
        from .authoring import StructureError

        structure, stem = self._cloze_structure()
        structure["cloze_blanks"][0]["options"][0]["is_correct"] = False
        with self.assertRaises(StructureError):
            self._validate(QuestionType.CLOZE, structure, stem=stem)

    def test_duplicate_blank_key_is_rejected(self):
        from .authoring import StructureError

        structure, stem = self._cloze_structure(stem="Assess [dropdown 1] and [DROPDOWN 1].")
        structure["cloze_blanks"].append(
            {
                "blank_key": "DROPDOWN 1",
                "display_order": 1,
                "options": [
                    {"option_text": "x", "is_correct": True, "rationale": ""},
                    {"option_text": "y", "is_correct": False, "rationale": ""},
                ],
            }
        )
        with self.assertRaises(StructureError):
            self._validate(QuestionType.CLOZE, structure, stem=stem)

    def test_valid_cloze_passes(self):
        structure, stem = self._cloze_structure()
        self._validate(QuestionType.CLOZE, structure, stem=stem)  # no exception

    # --- DRAG_DROP ---

    def test_dragdrop_item_with_unresolvable_category_key_is_rejected(self):
        from .authoring import StructureError

        structure = {
            "dragdrop_categories": [{"key": "cat0", "name": "Expected", "display_order": 0}],
            "dragdrop_items": [
                {
                    "text": "Item",
                    "display_order": 0,
                    "correct_category_key": "does-not-exist",
                    "correct_order": None,
                    "rationale": "",
                }
            ],
        }
        with self.assertRaises(StructureError):
            self._validate(QuestionType.DRAG_DROP, structure)

    def test_dragdrop_item_with_both_category_and_order_is_rejected(self):
        from .authoring import StructureError

        structure = {
            "dragdrop_categories": [{"key": "cat0", "name": "Expected", "display_order": 0}],
            "dragdrop_items": [
                {
                    "text": "Item",
                    "display_order": 0,
                    "correct_category_key": "cat0",
                    "correct_order": 1,
                    "rationale": "",
                }
            ],
        }
        with self.assertRaises(StructureError):
            self._validate(QuestionType.DRAG_DROP, structure)

    def test_dragdrop_sequencing_with_gap_in_order_is_rejected(self):
        from .authoring import StructureError

        structure = {
            "dragdrop_categories": [],
            "dragdrop_items": [
                {
                    "text": "Step 1",
                    "display_order": 0,
                    "correct_category_key": None,
                    "correct_order": 1,
                    "rationale": "",
                },
                {
                    "text": "Step 2",
                    "display_order": 1,
                    "correct_category_key": None,
                    "correct_order": 4,
                    "rationale": "",
                },
            ],
        }
        with self.assertRaises(StructureError):
            self._validate(QuestionType.DRAG_DROP, structure)

    def test_valid_dragdrop_category_variant_passes(self):
        structure = {
            "dragdrop_categories": [{"key": "cat0", "name": "Expected", "display_order": 0}],
            "dragdrop_items": [
                {
                    "text": "Item",
                    "display_order": 0,
                    "correct_category_key": "cat0",
                    "correct_order": None,
                    "rationale": "",
                }
            ],
        }
        self._validate(QuestionType.DRAG_DROP, structure)  # no exception

    def test_valid_dragdrop_sequencing_variant_passes(self):
        structure = {
            "dragdrop_categories": [],
            "dragdrop_items": [
                {
                    "text": "Step 1",
                    "display_order": 0,
                    "correct_category_key": None,
                    "correct_order": 1,
                    "rationale": "",
                },
                {
                    "text": "Step 2",
                    "display_order": 1,
                    "correct_category_key": None,
                    "correct_order": 2,
                    "rationale": "",
                },
            ],
        }
        self._validate(QuestionType.DRAG_DROP, structure)  # no exception

    # --- HOTSPOT ---

    def test_hotspot_target_absent_from_stem_is_rejected(self):
        from .authoring import StructureError

        structure = {
            "hotspot_targets": [
                {"target_text": "not in the stem", "is_correct": True, "display_order": 0, "rationale": ""}
            ]
        }
        with self.assertRaises(StructureError):
            self._validate(QuestionType.HOTSPOT, structure, stem="A client is diaphoretic and pale.")

    def test_hotspot_target_over_255_chars_is_rejected(self):
        from .authoring import StructureError

        long_text = "A" * 256
        structure = {
            "hotspot_targets": [
                {"target_text": long_text, "is_correct": True, "display_order": 0, "rationale": ""}
            ]
        }
        with self.assertRaises(StructureError):
            self._validate(QuestionType.HOTSPOT, structure, stem=long_text)

    def test_hotspot_with_zero_correct_targets_is_rejected(self):
        from .authoring import StructureError

        structure = {
            "hotspot_targets": [
                {"target_text": "diaphoretic", "is_correct": False, "display_order": 0, "rationale": ""}
            ]
        }
        with self.assertRaises(StructureError):
            self._validate(QuestionType.HOTSPOT, structure, stem="A client is diaphoretic and pale.")

    def test_valid_hotspot_passes(self):
        structure = {
            "hotspot_targets": [
                {"target_text": "diaphoretic", "is_correct": True, "display_order": 0, "rationale": ""}
            ]
        }
        self._validate(
            QuestionType.HOTSPOT, structure, stem="A client is diaphoretic and pale."
        )  # no exception


class NgnItemBankImporterDirectTests(TestCase):
    """
    Calls apps.questions.importer.NgnItemBankImporter directly (bypassing
    the management command entirely) — proves the extraction in Phase 6 of
    the implementation plan produces a genuinely reusable class, not just
    one that happens to still work through call_command.
    """

    def setUp(self):
        nursing_system, _ = NursingSystem.objects.get_or_create(name="Cardiovascular")
        Topic.objects.get_or_create(nursing_system=nursing_system, name="Heart Failure")
        Domain.objects.get_or_create(name="Adult Health")
        category, _ = ClientNeedsCategory.objects.get_or_create(name="Physiological Integrity")
        ClientNeedsSubcategory.objects.get_or_create(category=category, name="Reduction of Risk Potential")

    def _save_and_reopen(self, wb):
        import tempfile

        f = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
        wb.save(f.name)
        f.close()
        return f.name

    def test_run_returns_structured_result_matching_the_workbook(self):
        from apps.questions.importer import NgnItemBankImporter

        path = self._save_and_reopen(_build_workbook())
        result = NgnItemBankImporter().run(path)

        # 6 top-level questions (MCQ/MTX/BT/CLZ/EHS/DND) + 2 case-study items.
        self.assertEqual(result.created, 8)
        self.assertEqual(result.case_studies_created, 1)
        self.assertEqual(result.questions_imported, result.created + result.updated)
        self.assertEqual(result.rows_failed, 0)
        self.assertFalse(result.dry_run)

    def test_run_reports_row_errors_with_labels(self):
        from apps.questions.importer import NgnItemBankImporter

        wb = _build_workbook()
        ws = wb["Answer_Options"]
        ws.append(["EHS-1", "this text is not in the stem at all", "TRUE", "Ambiguous."])
        path = self._save_and_reopen(wb)

        result = NgnItemBankImporter().run(path)
        self.assertEqual(result.rows_failed, 1)
        self.assertEqual(result.errors[0].label, "EHS-1")
        self.assertIn("does not appear verbatim", result.errors[0].message)

    def test_dry_run_writes_nothing_and_result_says_so(self):
        from apps.questions.importer import NgnItemBankImporter

        path = self._save_and_reopen(_build_workbook())
        result = NgnItemBankImporter(dry_run=True).run(path)

        self.assertTrue(result.dry_run)
        self.assertEqual(Question.objects.count(), 0)
        self.assertGreater(result.created, 0)  # still reports what WOULD have been created


class WriteImportLogTests(TestCase):
    def setUp(self):
        nursing_system, _ = NursingSystem.objects.get_or_create(name="Cardiovascular")
        Topic.objects.get_or_create(nursing_system=nursing_system, name="Heart Failure")
        Domain.objects.get_or_create(name="Adult Health")
        category, _ = ClientNeedsCategory.objects.get_or_create(name="Physiological Integrity")
        ClientNeedsSubcategory.objects.get_or_create(category=category, name="Reduction of Risk Potential")

    def test_command_line_run_writes_a_log_with_null_uploader(self):
        from apps.questions.models import ImportLog

        with tempfile.NamedTemporaryFile(suffix=".xlsx") as f:
            _build_workbook().save(f.name)
            call_command("import_ngn_item_bank", file=f.name)

        log = ImportLog.objects.latest("uploaded_at")
        self.assertIsNone(log.uploaded_by)
        self.assertGreater(log.questions_imported, 0)
        self.assertEqual(log.rows_failed, 0)

    def test_dry_run_does_not_write_a_log(self):
        from apps.questions.models import ImportLog

        with tempfile.NamedTemporaryFile(suffix=".xlsx") as f:
            _build_workbook().save(f.name)
            call_command("import_ngn_item_bank", file=f.name, dry_run=True)

        self.assertEqual(ImportLog.objects.count(), 0)
