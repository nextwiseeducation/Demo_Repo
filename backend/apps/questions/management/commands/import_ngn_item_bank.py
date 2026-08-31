"""
Imports the full NGN Item Bank xlsx format: MCQ/SATA/EMR plus every
structural NGN type this project's schema supports (Matrix/Grid, Bow-Tie,
Drag-and-Drop, Cloze, Hot Spot, NGN Case Study).

Separate from import_choice_based_questions.py (JSON-format, MCQ/SATA/EMR
only) deliberately — that command's source file predates the xlsx template
and only ever needs to handle its original shape. This command's source
format (a 6-sheet workbook: Item_Master, Answer_Options, NGN_Cases,
NGN_Components, References, Valid_Values) is now the standard going
forward for every question type, including MCQ/SATA/EMR, since a future
batch has no reason to ever arrive as JSON again.
"""

import re
from pathlib import Path

import openpyxl
from django.core.exceptions import ValidationError
from django.core.management.base import BaseCommand, CommandError
from django.db import transaction

from apps.questions.models import (
    AnswerChoice,
    BowTieOption,
    BowTieSection,
    ClinicalJudgmentSkill,
    ClozeBlank,
    ClozeOption,
    CognitiveLevel,
    Difficulty,
    DragDropCategory,
    DragDropItem,
    HotSpotTarget,
    MatrixCell,
    MatrixColumn,
    MatrixRow,
    Question,
    QuestionType,
)
from apps.taxonomy.models import (
    CaseStudy,
    ClientNeedsCategory,
    ClientNeedsSubcategory,
    Domain,
    ExamType,
    NursingSystem,
    Subtopic,
    Tag,
    Topic,
)

# Item_Type -> QuestionType (or, for a case-study sub-item, -> ngn_type).
# Both Drag-and-Drop variants collapse to DRAG_DROP, matching how the model
# itself treats them as one type discriminated by which of
# DragDropItem.correct_category/correct_order is populated.
# "NGN Case Study" is deliberately absent — that Item_Master row is an
# index/summary only (see the template's own Read Me) and is never turned
# into a Question directly; its real content comes from NGN_Cases +
# NGN_Components instead.
ITEM_TYPE_MAP = {
    "MCQ": QuestionType.MCQ,
    "SATA": QuestionType.SATA,
    "EMR": QuestionType.EMR,
    "Matrix/Grid": QuestionType.MATRIX,
    "Bow-Tie": QuestionType.BOWTIE,
    "Drag-and-Drop (Sequencing)": QuestionType.DRAG_DROP,
    "Drag-and-Drop (Category Matching)": QuestionType.DRAG_DROP,
    "Cloze": QuestionType.CLOZE,
    "Hot Spot": QuestionType.HOTSPOT,
}

DIFFICULTY_MAP = {"Easy": Difficulty.EASY, "Moderate": Difficulty.MEDIUM, "Difficult": Difficulty.HARD}

COGNITIVE_LEVEL_MAP = {
    "Remember": CognitiveLevel.REMEMBER,
    "Understand": CognitiveLevel.UNDERSTAND,
    "Apply": CognitiveLevel.APPLY,
    "Analyze": CognitiveLevel.ANALYZE,
    "Evaluate": CognitiveLevel.EVALUATE,
    "Create": CognitiveLevel.CREATE,
}

CJ_SKILL_MAP = {
    "Recognize Cues": ClinicalJudgmentSkill.RECOGNIZE_CUES,
    "Analyze Cues": ClinicalJudgmentSkill.ANALYZE_CUES,
    "Prioritize Hypotheses": ClinicalJudgmentSkill.PRIORITIZE_HYPOTHESES,
    "Generate Solutions": ClinicalJudgmentSkill.GENERATE_SOLUTIONS,
    "Take Action": ClinicalJudgmentSkill.TAKE_ACTION,
    "Evaluate Outcomes": ClinicalJudgmentSkill.EVALUATE_OUTCOMES,
}

BOWTIE_SECTION_MAP = {
    "Action": BowTieSection.ACTION,
    "Condition": BowTieSection.CONDITION,
    "Assessment": BowTieSection.ASSESSMENT,
}


class RowError(Exception):
    """One row (or one whole case study) failed validation. Carries a human-readable reason. See import_choice_based_questions.py's identical RowError for the reasoning — rows are isolated and reported individually rather than aborting the whole run."""


def _is_true(value) -> bool:
    """Answer_Options' Is_Correct column stores the literal strings 'TRUE'/'FALSE', not real booleans."""
    return str(value).strip().upper() == "TRUE"


class Command(BaseCommand):
    """
    Imports every question type from the standard NGN Item Bank xlsx
    template (built this session — see Local/Question_Bank's own Read Me
    sheet for the full format spec).

    Unlike import_choice_based_questions.py, --file is REQUIRED rather than
    defaulting to a path inside the repo: the template lives outside
    Demo_Repo entirely (an editor's local working copy of client-supplied
    content, not a committed deploy artifact), so there is no single
    canonical path to default to, and each new batch is a genuinely
    different file. Run locally with DATABASE_URL pointed at whichever
    database (dev/production) should receive the import — same mechanism
    already used for the original 12-question batch.

    Idempotency: every top-level question is keyed by external_id =
    Question_ID. Every case-study sub-item is keyed by
    external_id = "{Case_ID}-item-{Item_No}". Every CaseStudy is keyed by
    external_id = Case_ID. Rows already present are left untouched by
    default; pass --update to refresh them from the file instead.

    A whole case study (its CaseStudy row + every linked NGN_Components
    item) is imported as ONE atomic unit, unlike top-level questions which
    are each independent — a case study with only some of its items
    imported is a broken case, so one bad item rolls the whole case back.

    KNOWN LIMITATION: NGN_Components' MCQ/SATA sub-items have their answer
    options embedded as inline "A) ... B) ... C) ..." text inside Stem, not
    as separable per-choice data the way top-level Item_Master rows are.
    There is nothing to build real AnswerChoice rows from without fragile
    text-parsing, so case-study sub-items import as reference/context
    content only — no gradeable choices, for any ngn_type. If a submission
    ever reaches one, grade_submission (apps.questions.services) already
    raises QuestionNotGradeable correctly rather than mis-grading.
    """

    help = "Imports all NGN Item Bank question types from the standard xlsx template."

    def add_arguments(self, parser):
        parser.add_argument("--file", required=True, help="Path to the NGN Item Bank xlsx file.")
        parser.add_argument(
            "--dry-run",
            action="store_true",
            help="Validate and report without writing anything. Use this to check a new content batch.",
        )
        parser.add_argument(
            "--update",
            action="store_true",
            help="Refresh questions/case studies that already exist (matched on external_id) instead of skipping them.",
        )

    def handle(self, *args, **options):
        path = Path(options["file"])
        if not path.exists():
            raise CommandError(f"{path} not found")
        dry_run = options["dry_run"]
        allow_update = options["update"]

        wb = openpyxl.load_workbook(path, data_only=True)
        sheets = self._read_sheets(wb)

        created = updated = skipped_existing = 0
        case_studies_created = case_studies_updated = 0
        errors: list[str] = []
        # See import_choice_based_questions.py's identical field for why
        # this is tracked and reported rather than silently allowed.
        created_taxonomy: list[str] = []

        for row in sheets["item_master"]:
            if row.get("Item_Type") == "NGN Case Study":
                # Index/summary row only — see this command's own docstring.
                continue
            label = row.get("Question_ID") or "(missing Question_ID)"
            try:
                with transaction.atomic():
                    outcome = self._import_question_row(row, sheets, allow_update, created_taxonomy)
                    if dry_run:
                        transaction.set_rollback(True)
                if outcome == "created":
                    created += 1
                elif outcome == "updated":
                    updated += 1
                else:
                    skipped_existing += 1
            except RowError as exc:
                errors.append(f"  {label}: {exc}")

        for case_id, case_row in sheets["ngn_cases"].items():
            try:
                with transaction.atomic():
                    case_outcome, item_outcomes = self._import_case_study(
                        case_row, sheets, allow_update, created_taxonomy
                    )
                    if dry_run:
                        transaction.set_rollback(True)
                if case_outcome == "created":
                    case_studies_created += 1
                elif case_outcome == "updated":
                    case_studies_updated += 1
                for outcome in item_outcomes:
                    if outcome == "created":
                        created += 1
                    elif outcome == "updated":
                        updated += 1
                    else:
                        skipped_existing += 1
            except RowError as exc:
                errors.append(f"  case {case_id}: {exc}")

        if dry_run:
            self.stdout.write(self.style.WARNING("DRY RUN — nothing was written."))

        self._report(created, updated, skipped_existing, case_studies_created, case_studies_updated, created_taxonomy, errors)

    # --- Sheet reading ---------------------------------------------------

    @staticmethod
    def _sheet_rows(wb, sheet_name):
        ws = wb[sheet_name]
        headers = [c.value for c in ws[1]]
        for row in ws.iter_rows(min_row=2, values_only=True):
            d = dict(zip(headers, row))
            if any(v is not None for v in d.values()):
                yield d

    def _read_sheets(self, wb):
        item_master = list(self._sheet_rows(wb, "Item_Master"))

        answer_options: dict[str, list[dict]] = {}
        for row in self._sheet_rows(wb, "Answer_Options"):
            if row.get("Question_ID"):
                answer_options.setdefault(row["Question_ID"], []).append(row)

        ngn_cases: dict[str, dict] = {}
        for row in self._sheet_rows(wb, "NGN_Cases"):
            if row.get("Case_ID"):
                ngn_cases[row["Case_ID"]] = row

        ngn_components: dict[str, list[dict]] = {}
        for row in self._sheet_rows(wb, "NGN_Components"):
            if row.get("Case_ID"):
                ngn_components.setdefault(row["Case_ID"], []).append(row)
        for rows in ngn_components.values():
            rows.sort(key=lambda r: r.get("Item_No") or 0)

        references: dict[str, str] = {}
        for row in self._sheet_rows(wb, "References"):
            if row.get("Reference_Key"):
                references[row["Reference_Key"]] = row.get("Full_Citation")

        return {
            "item_master": item_master,
            "answer_options": answer_options,
            "ngn_cases": ngn_cases,
            "ngn_components": ngn_components,
            "references": references,
        }

    # --- Top-level (non-case-study) questions -----------------------------

    def _import_question_row(self, row, sheets, allow_update, created_taxonomy) -> str:
        external_id = row.get("Question_ID")
        if not external_id:
            raise RowError("missing Question_ID — it is the key imports match on, so it is required")

        item_type_raw = row.get("Item_Type")
        question_type = ITEM_TYPE_MAP.get(item_type_raw)
        if question_type is None:
            raise RowError(f"unknown Item_Type {item_type_raw!r}")

        existing = Question.objects.filter(external_id=external_id).first()
        if existing is not None and not allow_update:
            return "skipped"

        domain = self._resolve_domain(row.get("Domain"), created_taxonomy)
        nursing_system, topic, subtopic = self._resolve_own_taxonomy(row, created_taxonomy)
        category, subcategory = self._resolve_client_needs(row)
        reference = self._resolve_reference(row.get("Reference_Key"), sheets)

        fields = {
            "external_id": external_id,
            "question_type": question_type,
            "stem": row.get("Stem") or "",
            "clinical_scenario": row.get("Scenario"),
            "difficulty": self._map_or_error(DIFFICULTY_MAP, row.get("Difficulty"), "Difficulty"),
            "domain": domain,
            "nursing_system": nursing_system,
            "topic": topic,
            "subtopic": subtopic,
            "nclex_client_needs_category": category,
            "nclex_client_needs_subcategory": subcategory,
            "clinical_judgment_skill": self._map_or_error(
                CJ_SKILL_MAP, row.get("Clinical_Judgment_Skill_Primary"), "Clinical_Judgment_Skill_Primary"
            ),
            "clinical_judgment_skill_secondary": CJ_SKILL_MAP.get(row.get("Clinical_Judgment_Skill_Secondary")),
            "cognitive_level": self._map_or_error(COGNITIVE_LEVEL_MAP, row.get("Cognitive_Level"), "Cognitive_Level"),
            "reference": reference,
            "key_takeaway": row.get("Clinical_Tip"),
        }

        question = existing or Question()
        for name, value in fields.items():
            setattr(question, name, value)

        try:
            question.full_clean(exclude=["id"])
        except ValidationError as exc:
            raise RowError(self._format_validation_error(exc)) from exc
        question.save()

        self._sync_tags(question, (row.get("Tags") or "").split(","), created_taxonomy)

        options = sheets["answer_options"].get(external_id, [])
        if question_type in (QuestionType.MCQ, QuestionType.SATA, QuestionType.EMR):
            self._sync_answer_choices(question, options)
        elif question_type == QuestionType.MATRIX:
            self._sync_matrix(question, options)
        elif question_type == QuestionType.BOWTIE:
            self._sync_bowtie(question, options)
        elif question_type == QuestionType.DRAG_DROP:
            if item_type_raw == "Drag-and-Drop (Sequencing)":
                self._sync_dragdrop_sequence(question, options)
            else:
                self._sync_dragdrop_category(question, options)
        elif question_type == QuestionType.CLOZE:
            self._sync_cloze(question, options)
        elif question_type == QuestionType.HOTSPOT:
            self._sync_hotspot(question, options)

        return "updated" if existing else "created"

    # --- NGN Case Study ----------------------------------------------------

    def _import_case_study(self, case_row, sheets, allow_update, created_taxonomy):
        case_id = case_row.get("Case_ID")
        if not case_id:
            raise RowError("missing Case_ID")

        existing_case = CaseStudy.objects.filter(external_id=case_id).first()
        if existing_case is not None and not allow_update:
            # Case already exists and no --update: still need to report its
            # existing items' status, but nothing to write.
            items = sheets["ngn_components"].get(case_id, [])
            return "skipped", ["skipped"] * len(items)

        case = existing_case or CaseStudy()
        case.external_id = case_id
        case.title = case_row.get("Topic/Subtopic") or ""
        case.shared_scenario = case_row.get("Case_Presentation_Hour0") or ""
        try:
            case.full_clean(exclude=["id"])
        except ValidationError as exc:
            raise RowError(self._format_validation_error(exc)) from exc
        case.save()
        case_outcome = "updated" if existing_case else "created"

        components = sheets["ngn_components"].get(case_id, [])
        if not components:
            raise RowError(f"no NGN_Components rows found for case {case_id}")

        item_outcomes = [
            self._import_case_item(case, case_row, comp, allow_update, sheets, created_taxonomy)
            for comp in components
        ]
        return case_outcome, item_outcomes

    def _import_case_item(self, case, case_row, comp, allow_update, sheets, created_taxonomy) -> str:
        item_no = comp.get("Item_No")
        if not item_no:
            raise RowError(f"case {case.external_id}: NGN_Components row missing Item_No")
        label = f"{case.external_id}-item-{item_no}"

        item_type_raw = comp.get("Item_Type")
        ngn_type = ITEM_TYPE_MAP.get(item_type_raw)
        if ngn_type is None:
            raise RowError(f"{label}: unknown Item_Type {item_type_raw!r}")

        external_id = f"{case.external_id}-item-{item_no}"
        existing = Question.objects.filter(external_id=external_id).first()
        if existing is not None and not allow_update:
            return "skipped"

        # Domain/Body_System/Topic/Reference_Key inherit from the case row
        # when the item doesn't override them — see this command's own
        # docstring and the template's Read Me. Client_Needs_Category has no
        # clean case-level equivalent to inherit from (NGN_Cases.Client_Needs
        # is a free-text combined description), so it's always required
        # directly on the item.
        taxonomy_row = {
            "Domain": comp.get("Domain") or case_row.get("Domain"),
            "Body_System": comp.get("Body_System") or case_row.get("Body_System"),
            "Topic": comp.get("Topic") or case_row.get("Topic/Subtopic"),
            "Subtopic": comp.get("Subtopic"),
            "Client_Needs_Category": comp.get("Client_Needs_Category"),
            "Client_Needs_Subcategory": comp.get("Client_Needs_Subcategory"),
        }
        domain = self._resolve_domain(taxonomy_row["Domain"], created_taxonomy)
        nursing_system, topic, subtopic = self._resolve_own_taxonomy(taxonomy_row, created_taxonomy)
        category, subcategory = self._resolve_client_needs(taxonomy_row)
        reference_key = comp.get("Reference_Key") or case_row.get("Reference_Key")
        reference = self._resolve_reference(reference_key, sheets)

        # Correct_Answer has no structured field to land in (see this
        # command's docstring on why case-study MCQ/SATA sub-items can't get
        # real AnswerChoice rows) — folded into rationale_correct instead of
        # being dropped.
        correct_answer = comp.get("Correct_Answer")
        rationale_text = comp.get("Rationale") or ""
        rationale_correct = f"Correct answer: {correct_answer}\n\n{rationale_text}" if correct_answer else rationale_text

        fields = {
            "external_id": external_id,
            "question_type": QuestionType.NGN_CASE,
            "ngn_type": ngn_type,
            "case_study": case,
            "case_study_sequence": item_no,
            "stem": comp.get("Stem") or "",
            "clinical_scenario": comp.get("Updated_Exhibit"),
            "difficulty": self._map_or_error(DIFFICULTY_MAP, comp.get("Difficulty"), "Difficulty", label),
            "domain": domain,
            "nursing_system": nursing_system,
            "topic": topic,
            "subtopic": subtopic,
            "nclex_client_needs_category": category,
            "nclex_client_needs_subcategory": subcategory,
            "clinical_judgment_skill": self._map_or_error(
                CJ_SKILL_MAP, comp.get("Clinical_Judgment_Step"), "Clinical_Judgment_Step", label
            ),
            "cognitive_level": self._map_or_error(COGNITIVE_LEVEL_MAP, comp.get("Cognitive_Level"), "Cognitive_Level", label),
            "reference": reference,
            "key_takeaway": comp.get("Clinical_Tip"),
            "rationale_correct": rationale_correct,
        }

        question = existing or Question()
        for name, value in fields.items():
            setattr(question, name, value)

        try:
            question.full_clean(exclude=["id"])
        except ValidationError as exc:
            raise RowError(f"{label}: {self._format_validation_error(exc)}") from exc
        question.save()

        # Opt-in: if the author has since added real Answer_Options rows for
        # this case item (keyed by its external_id, same as a top-level
        # question), sync them the same way _import_question_row does —
        # this is what upgrades a case-study sub-item from reference-only
        # content to something a student can actually answer and have
        # graded. Left un-synced (no RowError) when no rows are provided,
        # preserving this command's original "reference/context only"
        # default described in its own docstring above.
        options = sheets["answer_options"].get(external_id, [])
        if options:
            if ngn_type in (QuestionType.MCQ, QuestionType.SATA, QuestionType.EMR):
                self._sync_answer_choices(question, options)
            elif ngn_type == QuestionType.MATRIX:
                self._sync_matrix(question, options)
            elif ngn_type == QuestionType.BOWTIE:
                self._sync_bowtie(question, options)
            elif ngn_type == QuestionType.DRAG_DROP:
                if item_type_raw == "Drag-and-Drop (Sequencing)":
                    self._sync_dragdrop_sequence(question, options)
                else:
                    self._sync_dragdrop_category(question, options)
            elif ngn_type == QuestionType.CLOZE:
                self._sync_cloze(question, options)
            elif ngn_type == QuestionType.HOTSPOT:
                self._sync_hotspot(question, options)

        return "updated" if existing else "created"

    # --- Taxonomy resolution (shared by both paths above) -----------------

    @staticmethod
    def _resolve_domain(name, created_taxonomy):
        name = (name or "").strip()
        if not name:
            return None
        domain, made = Domain.objects.get_or_create(name=name)
        if made:
            created_taxonomy.append(f"Domain: {name}")
        return domain

    @staticmethod
    def _resolve_own_taxonomy(row, created_taxonomy):
        """NursingSystem/Topic/Subtopic — this project's own taxonomy, created on demand. See import_choice_based_questions.py's identical method for the full reasoning."""
        system_name = (row.get("Body_System") or "").strip()
        if not system_name:
            raise RowError("missing Body_System")
        nursing_system, made = NursingSystem.objects.get_or_create(name=system_name)
        if made:
            created_taxonomy.append(f"NursingSystem: {system_name}")

        topic_name = (row.get("Topic") or "").strip()
        if not topic_name:
            raise RowError("missing Topic")
        topic, made = Topic.objects.get_or_create(name=topic_name, nursing_system=nursing_system)
        if made:
            created_taxonomy.append(f"Topic: {system_name} / {topic_name}")

        subtopic = None
        subtopic_name = (row.get("Subtopic") or "").strip()
        if subtopic_name:
            subtopic, made = Subtopic.objects.get_or_create(name=subtopic_name, topic=topic)
            if made:
                created_taxonomy.append(f"Subtopic: {system_name} / {topic_name} / {subtopic_name}")

        return nursing_system, topic, subtopic

    @staticmethod
    def _resolve_client_needs(row):
        """Official NCSBN categories — looked up STRICTLY, never created. See import_choice_based_questions.py's identical method for the full reasoning."""
        category_name = (row.get("Client_Needs_Category") or "").strip()
        if not category_name:
            raise RowError("missing Client_Needs_Category")
        category = ClientNeedsCategory.objects.filter(name__iexact=category_name, exam_type=ExamType.RN).first()
        if category is None:
            raise RowError(f"unknown Client Needs category {category_name!r}")

        subcategory_name = (row.get("Client_Needs_Subcategory") or "").strip()
        if not subcategory_name:
            raise RowError("missing Client_Needs_Subcategory")
        subcategory = ClientNeedsSubcategory.objects.filter(name__iexact=subcategory_name, category=category).first()
        if subcategory is None:
            raise RowError(f"unknown Client Needs subcategory {subcategory_name!r} under {category_name!r}")

        return category, subcategory

    @staticmethod
    def _resolve_reference(key, sheets):
        key = (key or "").strip()
        if not key:
            return None
        citation = sheets["references"].get(key)
        if citation is None:
            raise RowError(f"unknown Reference_Key {key!r} — not found in the References sheet")
        return citation

    @staticmethod
    def _sync_tags(question, tag_names, created_taxonomy):
        tags = []
        for name in tag_names:
            name = (name or "").strip()
            if not name:
                continue
            tag, made = Tag.objects.get_or_create(name=name)
            if made:
                created_taxonomy.append(f"Tag: {name}")
            tags.append(tag)
        question.tags.set(tags)

    @staticmethod
    def _map_or_error(mapping, raw_value, field_name, label=None):
        value = mapping.get(raw_value)
        if value is None:
            prefix = f"{label}: " if label else ""
            raise RowError(f"{prefix}unknown {field_name} {raw_value!r}")
        return value

    # --- Per-type child-row builders ---------------------------------------

    @staticmethod
    def _sync_answer_choices(question, options):
        if not options:
            raise RowError("no Answer_Options rows for this question")
        correct_count = sum(1 for opt in options if _is_true(opt.get("Is_Correct")))
        if correct_count == 0:
            raise RowError("no answer choice is marked Is_Correct")
        if question.question_type == QuestionType.MCQ and correct_count > 1:
            raise RowError(f"question_type is MCQ but {correct_count} choices are marked correct")

        question.answer_choices.all().delete()
        for i, opt in enumerate(options):
            AnswerChoice.objects.create(
                question=question,
                choice_text=opt.get("Option_Text") or "",
                is_correct=_is_true(opt.get("Is_Correct")),
                display_order=i,
                rationale=opt.get("Rationale") or "",
            )

    @staticmethod
    def _sync_matrix(question, options):
        if not options:
            raise RowError("no Answer_Options rows for this Matrix/Grid question")
        column_names = list(dict.fromkeys(opt.get("Option_Text") for opt in options))
        if len(column_names) != 2:
            raise RowError(
                f"Matrix/Grid import only supports exactly 2 columns right now, found {len(column_names)}: "
                f"{column_names}"
            )

        question.matrix_rows.all().delete()
        question.matrix_columns.all().delete()
        columns = {
            name: MatrixColumn.objects.create(question=question, text=name, display_order=i)
            for i, name in enumerate(column_names)
        }
        for i, opt in enumerate(options):
            row_text = opt.get("Option_ID") or ""
            row = MatrixRow.objects.create(question=question, text=row_text, display_order=i)
            correct_col_name = opt.get("Option_Text")
            for name, col in columns.items():
                is_correct = name == correct_col_name
                MatrixCell.objects.create(
                    row=row, column=col, is_correct=is_correct,
                    rationale=(opt.get("Rationale") or "") if is_correct else "",
                )

    @staticmethod
    def _sync_bowtie(question, options):
        if not options:
            raise RowError("no Answer_Options rows for this Bow-Tie question")
        question.bowtie_options.all().delete()
        section_counters: dict[str, int] = {}
        for opt in options:
            option_id = opt.get("Option_ID") or ""
            prefix = option_id.split(" ")[0] if option_id else ""
            section = BOWTIE_SECTION_MAP.get(prefix)
            if section is None:
                raise RowError(f"Bow-Tie Option_ID {option_id!r} doesn't start with Action/Condition/Assessment")
            order = section_counters.get(section, 0)
            section_counters[section] = order + 1
            BowTieOption.objects.create(
                question=question, section=section, option_text=opt.get("Option_Text") or "",
                is_correct=_is_true(opt.get("Is_Correct")), display_order=order,
                rationale=opt.get("Rationale") or "",
            )

    @staticmethod
    def _sync_dragdrop_sequence(question, options):
        if not options:
            raise RowError("no Answer_Options rows for this Drag-and-Drop (Sequencing) question")
        question.dragdrop_items.all().delete()
        for opt in options:
            option_id = opt.get("Option_ID") or ""
            match = re.match(r"Step\s+(\d+)", option_id)
            if not match:
                raise RowError(f"Drag-and-Drop Option_ID {option_id!r} doesn't match 'Step N'")
            order = int(match.group(1))
            DragDropItem.objects.create(
                question=question, text=opt.get("Option_Text") or "", display_order=order - 1,
                correct_order=order, rationale=opt.get("Rationale") or "",
            )

    @staticmethod
    def _sync_dragdrop_category(question, options):
        if not options:
            raise RowError("no Answer_Options rows for this Drag-and-Drop (Category Matching) question")
        category_names = list(
            dict.fromkeys(opt.get("Option_Text") for opt in options if _is_true(opt.get("Is_Correct")))
        )
        if not category_names:
            raise RowError("no Is_Correct=TRUE rows to derive drag-drop categories from")

        question.dragdrop_items.all().delete()
        question.dragdrop_categories.all().delete()
        categories = {
            name: DragDropCategory.objects.create(question=question, name=name, display_order=i)
            for i, name in enumerate(category_names)
        }
        for i, opt in enumerate(options):
            is_correct = _is_true(opt.get("Is_Correct"))
            DragDropItem.objects.create(
                question=question, text=opt.get("Option_ID") or "", display_order=i,
                correct_category=categories.get(opt.get("Option_Text")) if is_correct else None,
                rationale=opt.get("Rationale") or "",
            )

    @staticmethod
    def _sync_cloze(question, options):
        if not options:
            raise RowError("no Answer_Options rows for this Cloze question")
        groups: dict[str, list[dict]] = {}
        for opt in options:
            option_id = opt.get("Option_ID") or ""
            match = re.match(r"Blank(\d+)_", option_id)
            blank_key = f"dropdown {match.group(1)}" if match else "dropdown 1"
            groups.setdefault(blank_key, []).append(opt)

        stem = question.stem or ""
        for blank_key in groups:
            if f"[{blank_key}]" not in stem:
                raise RowError(f"blank {blank_key!r} has answer options but no matching [{blank_key}] placeholder in Stem")

        question.cloze_blanks.all().delete()
        for i, blank_key in enumerate(sorted(groups, key=lambda k: int(k.split()[-1]))):
            blank = ClozeBlank.objects.create(question=question, blank_key=blank_key, display_order=i)
            for opt in groups[blank_key]:
                ClozeOption.objects.create(
                    blank=blank, option_text=opt.get("Option_Text") or "",
                    is_correct=_is_true(opt.get("Is_Correct")), rationale=opt.get("Rationale") or "",
                )

    @staticmethod
    def _sync_hotspot(question, options):
        if not options:
            raise RowError("no Answer_Options rows for this Hot Spot question")
        # A case-study item's own clinical_scenario is often just a short
        # hour label (e.g. "(Admission data, Hour 0)") — the actual passage
        # a Hot Spot item highlights within lives on the shared case, not
        # the item itself (Question.case_study's own comment). Both are
        # searched so a target can come from either.
        case_scenario = question.case_study.shared_scenario if question.case_study_id else ""
        haystack = (question.stem or "") + "\n" + (question.clinical_scenario or "") + "\n" + (case_scenario or "")
        question.hotspot_targets.all().delete()
        for i, opt in enumerate(options):
            target_text = opt.get("Option_Text") or ""
            if len(target_text) > 255:
                raise RowError(f"Hot Spot target text is {len(target_text)} chars, over the 255 limit: {target_text!r}")
            if target_text not in haystack:
                raise RowError(
                    f"Hot Spot target text {target_text!r} does not appear verbatim in the stem/scenario — "
                    "the highlighted phrase and the passage/table text must match exactly, word for word"
                )
            HotSpotTarget.objects.create(
                question=question, target_text=target_text, is_correct=_is_true(opt.get("Is_Correct")),
                display_order=i, rationale=opt.get("Rationale") or "",
            )

    # --- Reporting ----------------------------------------------------------

    @staticmethod
    def _format_validation_error(exc: ValidationError) -> str:
        if not hasattr(exc, "message_dict"):
            return "; ".join(exc.messages)
        return "; ".join(f"{field}: {' '.join(messages)}" for field, messages in exc.message_dict.items())

    def _report(self, created, updated, skipped_existing, case_studies_created, case_studies_updated, created_taxonomy, errors):
        self.stdout.write(self.style.SUCCESS(f"Imported {created} new question(s)."))
        if updated:
            self.stdout.write(self.style.SUCCESS(f"Updated {updated} existing question(s)."))
        if skipped_existing:
            self.stdout.write(
                f"Skipped {skipped_existing} question(s) that already exist (pass --update to refresh them)."
            )
        if case_studies_created:
            self.stdout.write(self.style.SUCCESS(f"Created {case_studies_created} new case stud(y/ies)."))
        if case_studies_updated:
            self.stdout.write(self.style.SUCCESS(f"Updated {case_studies_updated} existing case stud(y/ies)."))
        if created_taxonomy:
            distinct = list(dict.fromkeys(created_taxonomy))
            self.stdout.write(self.style.WARNING(f"Created {len(distinct)} new taxonomy row(s) — check for typos:"))
            for entry in distinct:
                self.stdout.write(f"  {entry}")
        if errors:
            self.stdout.write(self.style.ERROR(f"{len(errors)} row(s) failed validation:"))
            for error in errors:
                self.stdout.write(self.style.ERROR(error))
